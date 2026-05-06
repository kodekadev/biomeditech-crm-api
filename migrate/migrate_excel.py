"""
migrate_excel.py — Migra Cotizacion_base_con_historial4.xlsm a BigQuery.

Crea (si no existen) las tablas:
  - catalogo_servicios
  - plantillas_descripcion

Carga / actualiza (INSERT OR REPLACE equivalente via DELETE+INSERT):
  - catalogo_servicios   (164 items del catálogo)
  - plantillas_descripcion (10 textos base)
  - clientes             (93 clínicas del Excel)

Uso:
  set GOOGLE_APPLICATION_CREDENTIALS=C:/Users/bastian/Downloads/sa-key.json
  python migrate/migrate_excel.py
"""

import os, sys, uuid, re
from datetime import datetime, timezone
import openpyxl
from google.cloud import bigquery

# ── config ──────────────────────────────────────────────────────────────────

KEY = r"C:\Users\bastian\Downloads\sa-key.json"
if not os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = KEY

EXCEL = r"C:\Users\bastian\Downloads\Cotizacion_base_con_historial4.xlsm"
PROJECT = "jobs-425301"
DATASET = "CRM"
REF = f"`{PROJECT}.{DATASET}`"

client = bigquery.Client(project=PROJECT)

# ── helpers ──────────────────────────────────────────────────────────────────

def now_iso():
    return datetime.now(timezone.utc).isoformat()


def make_id(prefix: str) -> str:
    return f"{prefix}{uuid.uuid4().hex[:12].upper()}"


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def clean_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def run_sql(sql: str, job_config=None):
    job = client.query(sql, job_config=job_config)
    job.result()
    return job


# ── table creation ───────────────────────────────────────────────────────────

def add_column_if_missing():
    """Agrega descripcion_larga a servicios_cotizacion si aún no existe."""
    try:
        run_sql(f"ALTER TABLE {REF}.servicios_cotizacion ADD COLUMN IF NOT EXISTS descripcion_larga STRING")
        print("  ✓ servicios_cotizacion.descripcion_larga verificada")
    except Exception as e:
        print(f"  ⚠ ALTER TABLE: {e}")


def create_tables():
    ddls = [
        f"""
        CREATE TABLE IF NOT EXISTS {REF}.catalogo_servicios (
          id            STRING NOT NULL,
          codigo        STRING NOT NULL,
          categoria     STRING NOT NULL,
          servicio      STRING,
          equipo        STRING,
          unidad        STRING,
          precio_neto   INT64,
          grupo         STRING,
          texto_base_key STRING,
          activo        BOOL,
          creado_en     TIMESTAMP,
          _fecha_particion DATE
        )
        PARTITION BY _fecha_particion
        OPTIONS (require_partition_filter = false)
        """,
        f"""
        CREATE TABLE IF NOT EXISTS {REF}.plantillas_descripcion (
          id            STRING NOT NULL,
          codigo        STRING NOT NULL,
          descripcion_larga STRING,
          creado_en     TIMESTAMP,
          _fecha_particion DATE
        )
        PARTITION BY _fecha_particion
        OPTIONS (require_partition_filter = false)
        """,
    ]
    for ddl in ddls:
        print("  DDL:", ddl.strip()[:60], "...")
        run_sql(ddl)
    print("✓ Tablas verificadas / creadas")


# ── load helpers ─────────────────────────────────────────────────────────────

def truncate_and_load(table: str, rows: list[dict]):
    if not rows:
        print(f"  ⚠ sin filas para {table}")
        return
    run_sql(f"DELETE FROM {REF}.{table} WHERE TRUE")
    full = f"{PROJECT}.{DATASET}.{table}"
    errors = client.insert_rows_json(full, rows)
    if errors:
        print(f"  ✗ errores en {table}:", errors[:3])
    else:
        print(f"  ✓ {table}: {len(rows)} filas cargadas")


# ── migrate catalogo ─────────────────────────────────────────────────────────

def migrate_catalogo(wb):
    ws = wb["Lista de precio "]
    today = datetime.now(timezone.utc).isoformat()
    today_date = datetime.now(timezone.utc).date().isoformat()
    rows = []

    cat_map = {
        "VS": "VS", "MP": "MP", "MC": "MC",
        "BS": "BS", "EV": "EV", "RS": "RS",
    }

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        codigo = clean(row[0])
        if not codigo or len(codigo) < 2:
            continue
        cat_raw = clean(row[1]).upper()
        categoria = cat_map.get(cat_raw[:2], cat_raw[:2])
        if not categoria:
            continue

        precio = clean_int(row[5])
        rows.append({
            "id": make_id("CAT"),
            "codigo": codigo,
            "categoria": categoria,
            "servicio": clean(row[2]),
            "equipo": clean(row[3]),
            "unidad": clean(row[4]),
            "precio_neto": precio,
            "grupo": clean(row[6]),
            "texto_base_key": clean(row[7]),
            "activo": True,
            "creado_en": today,
            "_fecha_particion": today_date,
        })

    truncate_and_load("catalogo_servicios", rows)


# ── migrate plantillas ────────────────────────────────────────────────────────

def migrate_plantillas(wb):
    ws = wb["Descripciones"]
    today = datetime.now(timezone.utc).isoformat()
    today_date = datetime.now(timezone.utc).date().isoformat()
    rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        codigo = clean(row[0])
        desc = clean(row[1]) if row[1] else ""
        if not codigo or not desc:
            continue
        rows.append({
            "id": make_id("PLT"),
            "codigo": codigo,
            "descripcion_larga": desc,
            "creado_en": today,
            "_fecha_particion": today_date,
        })

    truncate_and_load("plantillas_descripcion", rows)


# ── migrate clientes ──────────────────────────────────────────────────────────

def migrate_clientes(wb):
    ws = wb["Datos cliente "]
    today = datetime.now(timezone.utc).isoformat()
    today_date = datetime.now(timezone.utc).date().isoformat()
    rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        nombre = clean(row[2])
        if not nombre or nombre == "None":
            continue

        tel_raw = clean(row[12])
        tel = re.sub(r"[^\d+]", "", tel_raw)[:20] if tel_raw else ""

        rows.append({
            "id": make_id("C"),
            "rut": clean(row[4]),
            "nombre_empresa": nombre,
            "contacto_nombre": clean(row[8]),
            "contacto_cargo": "",
            "email": clean(row[13]),
            "email_facturacion": "",
            "telefono": tel,
            "telefono_alt": "",
            "rubro": "Salud",
            "tipo_cliente": "clinica",
            "ciudad": clean(row[14]),
            "region": clean(row[1]),
            "comuna": clean(row[14]),
            "direccion": clean(row[6]),
            "estado": "activo",
            "creado_por": "migracion_excel",
            "creado_en": today,
            "_fecha_particion": today_date,
        })

    # Solo insertar si la tabla de clientes está vacía para no duplicar
    full = f"{PROJECT}.{DATASET}.clientes"
    count_job = client.query(f"SELECT COUNT(*) as n FROM {REF}.clientes").result()
    existing = list(count_job)[0]["n"]

    if existing > 0:
        print(f"  ⚠ clientes: ya hay {existing} registros, saltando migración (no queremos duplicar)")
        print(f"    Para forzar, ejecuta: DELETE FROM {REF}.clientes WHERE TRUE")
        return

    errors = client.insert_rows_json(full, rows)
    if errors:
        print(f"  ✗ errores en clientes:", errors[:3])
    else:
        print(f"  ✓ clientes: {len(rows)} filas cargadas")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=== Migracion Biomeditech Excel -> BigQuery ===\n")
    wb = openpyxl.load_workbook(EXCEL, read_only=True, keep_vba=False)

    print("1. Creando / actualizando tablas...")
    create_tables()
    add_column_if_missing()

    print("\n2. Cargando catálogo de servicios...")
    migrate_catalogo(wb)

    print("\n3. Cargando plantillas de descripción...")
    migrate_plantillas(wb)

    print("\n4. Cargando clientes...")
    migrate_clientes(wb)

    print("\n=== Migración completada ===")


if __name__ == "__main__":
    main()
